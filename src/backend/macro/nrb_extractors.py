import re
import unicodedata
import csv
from openpyxl import load_workbook
from pypdf import PdfReader
from .exceptions import NRBExtractionError

LABELS={
 "cpi_inflation":["overall inflation","national consumer price index"],
 "housing_inflation":["housing and utilities"],
 "lending_rate":["weighted average lending rate","lending rate commercial banks"],
 "deposit_rate":["weighted average deposit rate","deposit rate commercial banks"],
 "credit_growth":["private sector credit","credit to private sector"],
 "remittance_growth":["remittance inflows","remittance growth"],
}
def normalize(value):
    text=unicodedata.normalize("NFKD",str(value or "")).lower().replace("–","-").replace("—","-")
    return re.sub(r"\s+"," ",re.sub(r"[^a-z0-9.%+-]+"," ",text)).strip()
def numeric(value):
    if isinstance(value,(int,float)): return float(value)
    match=re.search(r"(?<!\d)-?\d+(?:\.\d+)?",normalize(value))
    return float(match.group()) if match else None

class NRBExcelExtractor:
    def extract(self,path):
        workbook=load_workbook(path,read_only=True,data_only=True); found={}; evidence={}
        for sheet in workbook.worksheets:
            rows=list(sheet.iter_rows(values_only=True))
            for r,row in enumerate(rows):
                joined=normalize(" ".join(str(v or "") for v in row))
                for field,labels in LABELS.items():
                    if field in found or not any(label in joined for label in labels): continue
                    candidates=[]
                    for rr in rows[r:min(len(rows),r+3)]: candidates.extend(numeric(v) for v in rr)
                    candidates=[v for v in candidates if v is not None and -100<v<500]
                    if candidates: found[field]=candidates[-1]; evidence[field]={"sheet":sheet.title,"source_label":joined[:500]}
        if len(found)<5: raise NRBExtractionError("Workbook did not contain a complete, unambiguous indicator set.")
        return found,evidence

class NRBPdfExtractor:
    def extract(self,path):
        text="\n".join(page.extract_text() or "" for page in PdfReader(path).pages); normalized=normalize(text); found={}; evidence={}
        for field,labels in LABELS.items():
            for label in labels:
                match=re.search(re.escape(label)+r".{0,220}?(-?\d+(?:\.\d+)?)\s*(?:percent|%)",normalized)
                if match: found[field]=float(match.group(1)); evidence[field]={"source_label":match.group(0)}; break
        if len(found)<5: raise NRBExtractionError("PDF did not contain a complete, unambiguous indicator set.")
        return found,evidence

class NRBCsvExtractor:
    def extract(self,path):
        with open(path,encoding="utf-8-sig",newline="") as handle: rows=list(csv.reader(handle))
        found={}; evidence={}
        for r,row in enumerate(rows):
            joined=normalize(" ".join(row))
            for field,labels in LABELS.items():
                if field in found or not any(label in joined for label in labels): continue
                candidates=[numeric(v) for rr in rows[r:min(len(rows),r+3)] for v in rr]
                candidates=[v for v in candidates if v is not None and -100<v<500]
                if candidates: found[field]=candidates[-1]; evidence[field]={"source_label":joined[:500]}
        if len(found)<5: raise NRBExtractionError("CSV did not contain a complete, unambiguous indicator set.")
        return found,evidence
